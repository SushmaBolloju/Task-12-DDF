import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service as ChromeService
import sys
import os
import time


class LoginPage:
    def __init__(self, driver):
        self.driver = driver
        self.url ="https://opensource-demo.orangehrmlive.com/web/index.php/auth/login"
        self.username_locator = (By.NAME,"username")
        self.password_locator = (By.NAME,"password")
        self.login_button_locator = (By.XPATH, "//button[@type='submit']")

    def navigate_to_login_page(self):
        self.driver.get(self.url)

    def enter_username(self, username):
        self.driver.find_element(*self.username_locator).send_keys("Admin")

    def enter_password(self, password):
        self.driver.find_element(*self.password_locator).send_keys("admin123")

    def click_login_button(self):
        self.driver.find_element(*self.login_button_locator).click()


sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

@pytest.fixture
def browser():
    chrome_service = ChromeService(r"C:\Users\chand\Downloads\chromedriver-win32\chromedriver-win32\chromedriver.exe")
    driver = webdriver.Chrome(service=chrome_service)
    driver.maximize_window()
    yield driver
    time.sleep(6)
    driver.quit()


def read_test_data_from_excel(file_path = r"D:\GuviTasks_Github\sheet1.xlsx",
                              sheet_name='Sheet1'):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    test_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:  # Ensure both username and password are present
            test_data.append((row[0], row[1]))
        else:
            print(f"Skipping invalid row: {row}")

    return test_data
def test_login(browser, username, password):
    login_page = LoginPage(browser)

    try:
        # Navigating to the  login page
        login_page.navigate_to_login_page()

        # Wait for username input field to be visible
        WebDriverWait(browser, 10).until(EC.visibility_of_element_located(login_page.username_locator))
        login_page.enter_username("Admin")


        # Wait for password input field to be visible
        WebDriverWait(browser, 10).until(EC.visibility_of_element_located(login_page.password_locator))
        login_page.enter_password("admin123")
        login_page.click_login_button()

        # Wait for login success
        WebDriverWait(browser, 10).until(EC.title_contains("Dashboard"))
    except Exception as e:
        print(f"Error during login process: {e}")
        raise


@pytest.mark.parametrize("username, password", read_test_data_from_excel())
def test_login_data_driven(browser, username, password):
    test_login(browser, username, password)
if __name__ == "__main__":
    pytest.main(["-v", "test_login.py"])